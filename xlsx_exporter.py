import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

def create_xlsx_file(data, file_path):
    workbook = Workbook()
    sheet = workbook.active

    # Заголовок
    headers = ["№", "Card Details"]
    sheet.append(headers)
    
    header_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, size=12, name='Calibri')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Данные
    alternating_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for index, item in enumerate(data, start=1):
        sheet.append([index, item])
        row = sheet[index + 1]
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        if index % 2 == 0:
            for cell in row:
                cell.fill = alternating_fill
    
    # Настройка ширины колонок
    sheet.column_dimensions['A'].width = 5
    sheet.column_dimensions['B'].width = 50

    # Стилизация границ
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    for row in sheet.iter_rows():
        for cell in row:
            cell.border = thin_border

    workbook.save(file_path)
